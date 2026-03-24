"""
rutas/reportes.py — Generación de reportes para el GAD.
Exporta Excel y PDF listos para presentar en el POA o Contraloría.
"""

import io
from pathlib import Path
from datetime import datetime, date
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))
from database import get_db
from models import Bache, Turno

router = APIRouter(prefix="/api/reportes", tags=["Reportes"])


def _filtrar_baches(db: Session, fecha_desde: Optional[date],
                    fecha_hasta: Optional[date], zona: Optional[str]):
    q = db.query(Bache)
    if fecha_desde:
        q = q.filter(Bache.fecha >= datetime.combine(fecha_desde, datetime.min.time()))
    if fecha_hasta:
        q = q.filter(Bache.fecha <= datetime.combine(fecha_hasta, datetime.max.time()))
    if zona:
        q = q.filter(Bache.zona.ilike(f"%{zona}%"))
    return q.order_by(Bache.fecha).all()


@router.get("/excel")
def exportar_excel(
    fecha_desde: Optional[date] = Query(None),
    fecha_hasta: Optional[date] = Query(None),
    zona:        Optional[str]  = Query(None),
    db: Session = Depends(get_db)
):
    """
    Exporta un Excel con todos los baches del período.
    Compatible con el formato de reporte del SENPLADES / POA.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return {"error": "Instala openpyxl: pip install openpyxl"}

    baches = _filtrar_baches(db, fecha_desde, fecha_hasta, zona)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventario de Baches"

    # ── Encabezado del municipio ─────────────────────────────────
    ws.merge_cells("A1:L1")
    ws["A1"] = "SISTEMA DE GESTIÓN VIAL MUNICIPAL"
    ws["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor="1A3A5C")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:L2")
    ws["A2"] = f"Reporte generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Período: {fecha_desde or 'todos'} → {fecha_hasta or 'todos'}"
    ws["A2"].font      = Font(size=10, color="FFFFFF")
    ws["A2"].fill      = PatternFill("solid", fgColor="2E6DA4")
    ws["A2"].alignment = Alignment(horizontal="center")

    # ── Cabecera de columnas ─────────────────────────────────────
    cabeceras = ["#", "Fecha", "Severidad", "Confianza %", "Latitud", "Longitud",
                 "Zona/Sector", "Estado", "Ancho (px)", "Alto (px)",
                 "Costo Est. USD", "Observaciones"]
    for col, cab in enumerate(cabeceras, 1):
        celda = ws.cell(row=3, column=col, value=cab)
        celda.font      = Font(bold=True, color="FFFFFF")
        celda.fill      = PatternFill("solid", fgColor="34495E")
        celda.alignment = Alignment(horizontal="center")

    colores_sev = {
        "critico":  "E74C3C", "grave":    "E67E22",
        "moderado": "F1C40F", "leve":     "2ECC71",
    }

    # ── Filas de datos ───────────────────────────────────────────
    for fila, b in enumerate(baches, 4):
        datos = [
            b.id,
            b.fecha.strftime("%d/%m/%Y %H:%M") if b.fecha else "",
            b.severidad.upper(),
            f"{b.confianza*100:.1f}%",
            f"{b.latitud:.6f}",
            f"{b.longitud:.6f}",
            b.zona or "—",
            b.estado.replace("_", " ").title(),
            b.ancho_px,
            b.alto_px,
            f"${b.costo_estimado:.2f}" if b.costo_estimado else "—",
            b.observaciones or "",
        ]
        for col, val in enumerate(datos, 1):
            c = ws.cell(row=fila, column=col, value=val)
            c.alignment = Alignment(horizontal="center")
            if col == 3:  # severidad con color
                color = colores_sev.get(b.severidad, "FFFFFF")
                c.fill = PatternFill("solid", fgColor=color)
                c.font = Font(bold=True, color="FFFFFF" if b.severidad in ("critico","grave") else "000000")
        if fila % 2 == 0:
            for col in range(1, 13):
                if ws.cell(row=fila, column=col).fill.fgColor.rgb == "00000000":
                    ws.cell(row=fila, column=col).fill = PatternFill("solid", fgColor="F2F2F2")

    # ── Resumen al final ─────────────────────────────────────────
    fila_res = len(baches) + 5
    ws.merge_cells(f"A{fila_res}:C{fila_res}")
    ws[f"A{fila_res}"] = "RESUMEN"
    ws[f"A{fila_res}"].font = Font(bold=True, size=11)

    costo_total = sum(b.costo_estimado or 0 for b in baches if b.estado != "reparado")
    resumen = [
        ("Total baches detectados", len(baches)),
        ("Baches reparados",        sum(1 for b in baches if b.estado == "reparado")),
        ("Baches pendientes",       sum(1 for b in baches if b.estado != "reparado")),
        ("Baches críticos activos", sum(1 for b in baches if b.severidad == "critico" and b.estado != "reparado")),
        ("Costo estimado total USD",f"${costo_total:.2f}"),
    ]
    for i, (etiqueta, valor) in enumerate(resumen, fila_res + 1):
        ws.cell(row=i, column=1, value=etiqueta).font = Font(bold=True)
        ws.cell(row=i, column=4, value=valor)

    # Ancho de columnas
    anchos = [5, 16, 12, 12, 12, 12, 20, 16, 10, 10, 14, 30]
    for i, a in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = a

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nombre_archivo = f"reporte_baches_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nombre_archivo}"}
    )


@router.get("/resumen-poa")
def resumen_poa(
    fecha_desde: Optional[date] = Query(None),
    fecha_hasta: Optional[date] = Query(None),
    db: Session = Depends(get_db)
):
    """
    Resumen ejecutivo en JSON — para copiar en el informe POA.
    Incluye porcentaje de cumplimiento de meta de reparación.
    """
    baches = _filtrar_baches(db, fecha_desde, fecha_hasta, None)
    total  = len(baches)
    if total == 0:
        return {"mensaje": "No hay datos en el período seleccionado"}

    reparados  = sum(1 for b in baches if b.estado == "reparado")
    costo_est  = sum(b.costo_estimado or 0 for b in baches)
    por_sev    = {}
    por_zona   = {}
    for b in baches:
        por_sev[b.severidad]  = por_sev.get(b.severidad, 0) + 1
        zona = b.zona or "Sin zona"
        por_zona[zona] = por_zona.get(zona, 0) + 1

    # Zona más crítica
    zona_critica = max(por_zona, key=por_zona.get) if por_zona else "N/D"

    return {
        "periodo": {
            "desde": str(fecha_desde) if fecha_desde else "inicio",
            "hasta": str(fecha_hasta) if fecha_hasta else "hoy",
        },
        "indicadores": {
            "total_baches_detectados":  total,
            "baches_reparados":         reparados,
            "baches_pendientes":        total - reparados,
            "porcentaje_resolucion":    f"{reparados/total*100:.1f}%",
            "costo_estimado_total_usd": round(costo_est, 2),
            "zona_mas_afectada":        zona_critica,
        },
        "distribucion_severidad": por_sev,
        "distribucion_por_zona":  por_zona,
        "meta_sugerida": f"Reparar el 100% de baches críticos ({por_sev.get('critico', 0)}) en 30 días",
    }
